#!/usr/bin/env python3
"""
DOP Agent Portal — Deposit Accounts Automation Script
=====================================================
Site: dopagent.indiapost.gov.in  (Finacle Agent Portal)
Input: XLSX with columns — LOT, RD Numbers (comma-separated), Count, Timestamp
Output: Updated XLSX (in-place) + downloaded PDFs

COMPLETE FLOW
─────────────
 STARTUP
   1. Read LOTs from XLSX (each LOT = a batch of RD account numbers)
   2. Open Chrome (download dir pre-set to ~/Downloads/LOT_<today>/)
   3. User logs in manually → navigates to Deposit Accounts page → presses ENTER

 PHASE 1 — For each LOT (on the Deposit Accounts page):
   Step 1.  Ensure "Cash" radio is selected (skip click if already selected)
   Step 2.  Clear textarea with Cmd+A → Delete (avoids page refresh)
            Enter the LOT's RD numbers into the Account Id(s) textarea
   Step 3.  Click "Fetch" → wait for results table
   Step 4.  Read "Displaying 1 - X of Y results" → verify Y == expected Count
   Step 5.  Validate every row's "Next RD Installment Due Date" is in current month
            → If any mismatch → SKIP this LOT (log bad accounts)
   Step 6.  Select all checkboxes across all pages (pagination-safe)
   Step 7.  Re-read display count → verify selected == total ("X of X")
   Step 8.  Click "Save" → portal redirects to "Selected Recurring Deposit Account List"
   Step 9.  Click "Pay All Saved Installments"
            → Parse success message for Reference ID (e.g. C320461082)
            → Store Reference_ID in XLSX (green column)
            → Portal auto-redirects back to Deposit Accounts page
   Step 10. Save progress to XLSX after every LOT
            → Next LOT starts from Step 1 (old text cleared, Cash stays selected)

 PHASE 2 — PDF Downloads (after all LOTs are done):
   1. Navigate to Reports → "Recurring Deposit Installment Report"
   2. For each LOT that has a Reference_ID:
      a. Enter Reference ID in "List Reference No" field
      b. Click "Search" → verify result count matches expected Count
      c. Ensure "PDF file" is selected in dropdown → click "OK"
      d. Wait for download → rename file to <LOT#>_<RefID>.pdf
      e. Manually clear reference field (Cmd+A → Delete) → repeat for next LOT
         (does NOT click Clear button — that would reset date fields too)
   3. All PDFs saved to: ~/Downloads/LOT_<YYYY-MM-DD>/

 XLSX COLUMNS (written after each LOT):
   LOT | RD Numbers | Count | Reference_ID | Timestamp
   Fetch_Status | Count_Match | Due_Date_Check | Selected
   Selection_Verified | Save_Status | Pay_Status | Remarks

 RESUMABILITY:
   - LOTs with Pay_Status=OK are skipped on re-run
   - PDFs already on disk (matching filename) are skipped
   - Progress saved to XLSX after every single LOT

 PACING:
   All actions have deliberate delays (DELAY_SHORT=1.5s, DELAY_MEDIUM=3s,
   DELAY_LONG=5s, DELAY_CHECKBOX=0.4s) to avoid spam-like behaviour on the
   banking portal. No confirmation prompts between LOTs — just a steady pace.
"""

import time
import sys
import re
import os
import glob as glob_mod
import platform
import threading
import psutil
from datetime import datetime, date
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from dop_dashboard import (
    DashboardState, ControlFlags, SkipLotException,
    StopAfterCurrentException, checkpoint, start_dashboard
)

# ── Dashboard shared state ──
dashboard_state = DashboardState()
control_flags = ControlFlags()

# Override print to also feed dashboard log
_original_print = print
def print(*args, **kwargs):
    sep = kwargs.get("sep", " ")
    msg = sep.join(str(a) for a in args)
    _original_print(*args, **kwargs)
    timestamp = datetime.now().strftime("%H:%M:%S")
    with dashboard_state.lock:
        dashboard_state.log_messages.append(f"{timestamp}  {msg}")

# ── Configuration ──
XLSX_FILE = ""  # Set at startup from user-provided Excel path (primary input/output)
PORTAL_URL = "https://dopagent.indiapost.gov.in/corp/Finacle"
WAIT_TIMEOUT = 30  # seconds to wait for elements

# Delays (seconds) - kept gentle to avoid spam-like behaviour
DELAY_SHORT = 1.5     # after small actions (clicking radio, clearing fields)
DELAY_MEDIUM = 3.0    # after fetch, page loads
DELAY_LONG = 5.0      # between LOTs, after save
DELAY_CHECKBOX = 0.4  # between each checkbox click
GLOBAL_TIMEOUT_MINS = 30  # Auto-exit after this many minutes to prevent hangs

# Current month for due-date validation (e.g. "Feb" for February)
CURRENT_MONTH_ABBR = datetime.now().strftime("%b")  # "Feb", "Mar", etc.

# Key modifier for select-all: Cmd on Mac, Ctrl on others
SELECT_ALL_KEY = Keys.COMMAND if platform.system() == "Darwin" else Keys.CONTROL

# Download folder for Phase 2 PDFs
TODAY_STR = date.today().strftime("%Y-%m-%d")  # e.g. "2026-02-21"
DOWNLOAD_DIR = os.path.expanduser(f"~/Downloads/LOT_{TODAY_STR}")

# XLSX columns — Reference_ID is right after Count
XLSX_COLUMNS = [
    "LOT", "RD Numbers", "Count", "Reference_ID", "Timestamp",
    "Fetch_Status", "Count_Match", "Due_Date_Check",
    "Selected", "Selection_Verified", "Save_Status",
    "Pay_Status", "Remarks"
]


# ── XLSX Read / Write ──

def read_xlsx(filepath):
    """Read the Excel file and return list of dicts with all columns."""
    from openpyxl import load_workbook
    wb = load_workbook(filepath)
    ws = wb.active
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    lots = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue  # skip fully empty rows
        rd = dict(zip(headers, row))
        lot = {
            "LOT":                str(rd.get("LOT", "") or "").strip(),
            "RD Numbers":         str(rd.get("RD Numbers", "") or "").strip(),
            "Count":              int(rd.get("Count", 0) or 0),
            "Reference_ID":       str(rd.get("Reference_ID", "") or "").strip(),
            "Timestamp":          str(rd.get("Timestamp", "") or "").strip(),
            "Fetch_Status":       str(rd.get("Fetch_Status", "") or "").strip(),
            "Count_Match":        str(rd.get("Count_Match", "") or "").strip(),
            "Due_Date_Check":     str(rd.get("Due_Date_Check", "") or "").strip(),
            "Selected":           str(rd.get("Selected", "") or "").strip(),
            "Selection_Verified": str(rd.get("Selection_Verified", "") or "").strip(),
            "Save_Status":        str(rd.get("Save_Status", "") or "").strip(),
            "Pay_Status":         str(rd.get("Pay_Status", "") or "").strip(),
            "Remarks":            str(rd.get("Remarks", "") or "").strip(),
        }
        lots.append(lot)
    return lots


def write_xlsx(filepath, lots):
    """Write a formatted XLSX with green Reference_ID column."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "RD Session"

        # Header row
        headers = XLSX_COLUMNS
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Green fill for Reference_ID column
        green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        green_font = Font(bold=True, size=11)
        ref_col_idx = headers.index("Reference_ID") + 1  # 1-based

        # Data rows
        for row_idx, lot in enumerate(lots, 2):
            for col_idx, header in enumerate(headers, 1):
                value = lot.get(header, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Green highlight for Reference_ID column
                if col_idx == ref_col_idx and value:
                    cell.fill = green_fill
                    cell.font = green_font

        # Auto-fit column widths (approximate)
        col_widths = {
            "LOT": 6, "RD Numbers": 60, "Count": 7, "Reference_ID": 18,
            "Timestamp": 22, "Fetch_Status": 13, "Count_Match": 18,
            "Due_Date_Check": 16, "Selected": 10, "Selection_Verified": 20,
            "Save_Status": 13, "Pay_Status": 12, "Remarks": 40
        }
        for col_idx, header in enumerate(headers, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = col_widths.get(header, 15)

        # Freeze top row
        ws.freeze_panes = "A2"

        wb.save(filepath)
        print(f"  Formatted XLSX saved: {filepath}")
    except ImportError:
        print("  ⚠ openpyxl not installed, skipping XLSX generation")
    except Exception as e:
        print(f"  ⚠ Could not write XLSX: {e}")


# ── Global timeout ──

def _global_timeout_handler(lots):
    """Called by the timer thread when the global timeout expires."""
    print(f"\n{'=' * 60}")
    print(f"  GLOBAL TIMEOUT ({GLOBAL_TIMEOUT_MINS} min) REACHED")
    print(f"  Saving progress and exiting to prevent system hang.")
    print(f"{'=' * 60}")
    try:
        write_xlsx(XLSX_FILE, lots)
        print(f"  Progress saved. Restart the script to resume.")
    except Exception as e:
        print(f"  ⚠ Could not save progress: {e}")
    os._exit(1)  # Hard exit — works even if main thread is stuck in input()/sleep()


def start_global_timeout(lots):
    """Start a daemon timer that will force-exit after GLOBAL_TIMEOUT_MINS."""
    seconds = GLOBAL_TIMEOUT_MINS * 60
    timer = threading.Timer(seconds, _global_timeout_handler, args=[lots])
    timer.daemon = True  # Won't prevent normal exit
    timer.start()
    print(f"  Global timeout set: {GLOBAL_TIMEOUT_MINS} minutes")
    return timer


# ── Memory watchdog ──

MEMORY_LIMIT_MB = 3500  # Kill script if Chrome + Python exceed this


def _get_memory_mb():
    """Get total memory usage (Python + Chrome automation) in MB."""
    try:
        python_proc = psutil.Process(os.getpid())
        python_mb = python_proc.memory_info().rss / (1024 * 1024)
        chrome_mb = 0
        for proc in psutil.process_iter(['pid', 'name', 'cmdline']):
            try:
                if proc.info['name'] and 'chrome' in proc.info['name'].lower():
                    cmdline = proc.info.get('cmdline') or []
                    if any('--test-type=webdriver' in arg for arg in cmdline):
                        chrome_mb += proc.memory_info().rss / (1024 * 1024)
            except (psutil.NoSuchProcess, psutil.AccessDenied):
                continue
        return python_mb + chrome_mb
    except Exception:
        return 0.0


def check_memory_usage(driver):
    """Check total memory used by Chrome + Python. Warn or abort if too high."""
    total_mb = _get_memory_mb()
    with dashboard_state.lock:
        dashboard_state.memory_mb = total_mb
    if total_mb > MEMORY_LIMIT_MB:
        print(f"\n⚠ MEMORY WARNING: {total_mb:.0f} MB used (limit: {MEMORY_LIMIT_MB} MB)")
        print(f"  Saving progress and stopping to prevent system freeze.")
        return False
    return True


# ── Browser helpers ──

def kill_previous_automation_chrome():
    """Kill Chrome instances from previous automation runs (webdriver-spawned only)."""
    killed = 0
    for proc in psutil.process_iter(['pid', 'name', 'cmdline']):
        try:
            if proc.info['name'] and 'chrome' in proc.info['name'].lower():
                cmdline = proc.info.get('cmdline') or []
                # Only kill Chrome instances launched by WebDriver (have --test-type=webdriver flag)
                if any('--test-type=webdriver' in arg for arg in cmdline):
                    proc.kill()
                    killed += 1
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            continue
    if killed:
        print(f"  Cleaned up {killed} leftover automation Chrome process(es)")
        time.sleep(2)  # Give OS time to reclaim memory
    else:
        print(f"  No leftover automation Chrome processes found")


def setup_driver(download_dir=None):
    """Create and return a Chrome WebDriver with optional download directory."""
    # Kill any leftover Chrome from previous automation runs
    kill_previous_automation_chrome()

    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-extensions")
    options.add_experimental_option("detach", True)

    if download_dir:
        os.makedirs(download_dir, exist_ok=True)
        prefs = {
            "download.default_directory": download_dir,
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "plugins.always_open_pdf_externally": True,
        }
        options.add_experimental_option("prefs", prefs)

    driver = webdriver.Chrome(options=options)
    return driver


def wait_for_login(driver):
    """Open the portal and wait for the user to log in manually."""
    driver.get(PORTAL_URL)
    print("\n" + "=" * 60)
    print("  PLEASE LOG IN TO THE DOP AGENT PORTAL MANUALLY")
    print("=" * 60)
    print("\nAfter logging in, navigate to:")
    print("  Accounts → Agent Enquire & Update Screen → Deposit Accounts")
    print("\nOnce you are on the Deposit Accounts page, press ENTER here...")
    input()
    print("Continuing with automation...\n")


def navigate_to_deposit_accounts(driver, wait):
    """Verify we are on the Deposit Accounts page."""
    # Small pause to let page fully render after user presses ENTER
    time.sleep(dashboard_state.delay_short)
    # Check page title/heading — use text() to avoid matching every ancestor element
    try:
        driver.find_element(By.XPATH,
            "//b[contains(text(),'DEPOSIT ACCOUNTS')] | //h1[contains(text(),'DEPOSIT ACCOUNTS')] | //h2[contains(text(),'DEPOSIT ACCOUNTS')] | //span[contains(text(),'DEPOSIT ACCOUNTS')] | //td[contains(text(),'DEPOSIT ACCOUNTS')]")
        print("✓ On Deposit Accounts page")
        return True
    except NoSuchElementException:
        # Fallback: check for the Fetch button or textarea (unique to this page)
        try:
            driver.find_element(By.XPATH,
                "//input[@value='Fetch'] | //textarea")
            print("✓ On Deposit Accounts page (detected via Fetch button)")
            return True
        except NoSuchElementException:
            print("⚠ Could not auto-detect Deposit Accounts page.")
            print("  (This is OK — continuing since you confirmed with ENTER)")
            return True


def ensure_cash_mode(driver):
    """Select Cash radio only if not already selected (avoids unnecessary clicks)."""
    try:
        cash_radio = driver.find_element(By.XPATH, "//input[@type='radio' and @value='C']")
        if cash_radio.is_selected():
            print("  ✓ Cash mode already selected")
            return
        cash_radio.click()
        time.sleep(dashboard_state.delay_short)
        print("  ✓ Cash mode selected")
    except NoSuchElementException:
        try:
            radios = driver.find_elements(By.XPATH, "//input[@type='radio']")
            if radios and not radios[0].is_selected():
                radios[0].click()
                time.sleep(dashboard_state.delay_short)
            print("  ✓ Cash mode selected (first radio)")
        except Exception as e:
            print(f"  ⚠ Could not select Cash mode: {e}")


def clear_textarea_and_enter(driver, rd_numbers):
    """
    Clear the textarea by selecting all text and deleting it (no page refresh),
    then type in the new RD numbers.
    """
    textarea = None
    try:
        textarea = driver.find_element(By.TAG_NAME, "textarea")
    except NoSuchElementException:
        try:
            textarea = driver.find_element(By.XPATH,
                "//textarea | //input[contains(@name,'account') or contains(@name,'Account')]")
        except NoSuchElementException:
            print("  ✗ Could not find Account ID input field!")
            return False

    # Select all existing text and delete (no page refresh unlike Clear Account btn)
    textarea.click()
    time.sleep(dashboard_state.delay_checkbox)
    textarea.send_keys(SELECT_ALL_KEY, "a")
    time.sleep(dashboard_state.delay_checkbox)
    textarea.send_keys(Keys.DELETE)
    time.sleep(dashboard_state.delay_short)

    # Type new RD numbers
    textarea.send_keys(rd_numbers)
    time.sleep(dashboard_state.delay_short)
    print("  ✓ Cleared old text & entered new RD numbers")
    return True


def click_fetch(driver, wait):
    """Click the Fetch button and wait for results."""
    try:
        fetch_btn = driver.find_element(By.XPATH, "//input[@value='Fetch' or contains(@value,'Fetch')]")
        fetch_btn.click()
        print("  ✓ Clicked Fetch")
        time.sleep(dashboard_state.delay_medium)
        try:
            wait.until(EC.presence_of_element_located(
                (By.XPATH, "//*[contains(text(), 'Displaying')]")
            ))
        except TimeoutException:
            time.sleep(dashboard_state.delay_medium)
        return True
    except NoSuchElementException:
        print("  ✗ Could not find Fetch button!")
        return False


# ── Display count parsing ──

def get_display_text(driver):
    """Get the raw 'Displaying X - Y of Z result(s)' text."""
    try:
        el = driver.find_element(By.XPATH, "//*[contains(text(), 'Displaying')]")
        return el.text.strip()
    except NoSuchElementException:
        return None


def parse_display_count(text):
    """Parse 'Displaying 1 - 7 of 7 results' → (1, 7, 7)."""
    if not text:
        return None
    m = re.search(r'Displaying\s+(\d+)\s*-\s*(\d+)\s+of\s+(\d+)', text)
    if m:
        return int(m.group(1)), int(m.group(2)), int(m.group(3))
    return None


# ── Pagination helpers (targets pagination area, NOT column headers) ──

def find_next_page_button(driver):
    """Find the '>' pagination button near 'Page X of Y', not column sort links."""
    try:
        page_info = driver.find_element(By.XPATH,
            "//*[contains(text(), 'Page') and contains(text(), 'of')]")
        parent = page_info.find_element(By.XPATH, "./..")
        try:
            return parent.find_element(By.XPATH, ".//a[normalize-space(text())='>']")
        except NoSuchElementException:
            grandparent = parent.find_element(By.XPATH, "./..")
            return grandparent.find_element(By.XPATH, ".//a[normalize-space(text())='>']")
    except NoSuchElementException:
        return None


def find_prev_page_button(driver):
    """Find the '<' pagination button in the pagination area."""
    try:
        page_info = driver.find_element(By.XPATH,
            "//*[contains(text(), 'Page') and contains(text(), 'of')]")
        parent = page_info.find_element(By.XPATH, "./..")
        try:
            return parent.find_element(By.XPATH, ".//a[normalize-space(text())='<']")
        except NoSuchElementException:
            grandparent = parent.find_element(By.XPATH, "./..")
            return grandparent.find_element(By.XPATH, ".//a[normalize-space(text())='<']")
    except NoSuchElementException:
        return None


def go_to_page_1(driver):
    """Navigate back to page 1 (capped at 10 clicks to prevent runaway loop)."""
    for _ in range(10):
        prev_btn = find_prev_page_button(driver)
        if prev_btn:
            try:
                prev_btn.click()
                time.sleep(dashboard_state.delay_medium)
            except Exception:
                # Transient error (stale element, etc.) — re-find and retry
                continue
        else:
            break  # No prev button means we are on page 1
    # Fallback: Go to Page input
    try:
        page_input = driver.find_element(By.XPATH,
            "//input[@type='text' and (contains(@name,'page') or contains(@title,'Page'))]")
        page_input.clear()
        page_input.send_keys("1")
        go_btn = driver.find_element(By.XPATH, "//input[@value='Go']")
        go_btn.click()
        time.sleep(dashboard_state.delay_medium)
    except NoSuchElementException:
        pass


def get_total_pages(driver):
    """Parse 'Page 1 of 16' and return total pages."""
    try:
        el = driver.find_element(By.XPATH,
            "//*[contains(text(), 'Page') and contains(text(), 'of')]")
        m = re.search(r'Page\s+\d+\s+of\s+(\d+)', el.text.strip())
        if m:
            return int(m.group(1))
    except NoSuchElementException:
        pass
    return 1


# ── Due date validation ──

def validate_due_dates_on_page(driver):
    """Check due dates on current page. Returns list of (account_no, due_date) for bad rows."""
    bad_rows = []
    rows = driver.find_elements(By.XPATH, "//table//tr[td]")
    for row in rows:
        cells = row.find_elements(By.TAG_NAME, "td")
        if len(cells) >= 6:
            account_no = cells[1].text.strip()
            due_date = cells[5].text.strip()
            if due_date and CURRENT_MONTH_ABBR not in due_date:
                bad_rows.append((account_no, due_date))
    return bad_rows


def validate_due_dates_all_pages(driver, expected_count=0):
    """Validate due dates across all pages, then return to page 1.
    Skips pagination entirely when expected_count <= 10 (single page)."""
    all_bad = []

    all_bad.extend(validate_due_dates_on_page(driver))

    # Only paginate if there are more than 10 accounts (i.e. multiple pages)
    if expected_count > 10:
        total_pages = get_total_pages(driver)
        for _ in range(total_pages - 1):
            next_btn = find_next_page_button(driver)
            if next_btn:
                next_btn.click()
                time.sleep(dashboard_state.delay_medium)
                all_bad.extend(validate_due_dates_on_page(driver))
            else:
                break

        if total_pages > 1:
            go_to_page_1(driver)

    return all_bad


# ── Checkbox selection ──

def select_all_checkboxes_on_page(driver):
    """Select all checkboxes in data rows (inside <td>, not <th>)."""
    checkboxes = driver.find_elements(By.XPATH, "//table//td//input[@type='checkbox']")
    selected = 0
    for cb in checkboxes:
        try:
            if not cb.is_selected():
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", cb)
                time.sleep(dashboard_state.delay_checkbox)
                cb.click()
                time.sleep(dashboard_state.delay_checkbox)
            selected += 1
        except Exception as e:
            print(f"    ⚠ Could not click a checkbox: {e}")
    print(f"    ✓ Selected {selected} checkboxes on this page")
    return selected


def select_all_checkboxes_all_pages(driver, expected_count=0):
    """Select checkboxes across all pages, returns total selected.
    Skips pagination entirely when expected_count <= 10 (single page)."""
    total_selected = 0

    total_selected += select_all_checkboxes_on_page(driver)

    # Only paginate if there are more than 10 accounts (i.e. multiple pages)
    if expected_count > 10:
        total_pages = get_total_pages(driver)
        for _ in range(total_pages - 1):
            next_btn = find_next_page_button(driver)
            if next_btn:
                next_btn.click()
                time.sleep(dashboard_state.delay_medium)
                total_selected += select_all_checkboxes_on_page(driver)
            else:
                break

        # Return to page 1 so Save button is in the correct page context
        if total_pages > 1:
            go_to_page_1(driver)

    return total_selected


# ── Save ──

def click_save(driver, wait):
    """Click the Save button on the Deposit Accounts page."""
    try:
        save_btn = driver.find_element(By.XPATH, "//input[@value='Save' or contains(@value,'Save')]")
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", save_btn)
        time.sleep(dashboard_state.delay_short)
        save_btn.click()
        print("  ✓ Clicked Save")
        time.sleep(dashboard_state.delay_long)

        # Handle any confirmation alert
        try:
            alert = driver.switch_to.alert
            print(f"    Alert: {alert.text}")
            alert.accept()
            time.sleep(dashboard_state.delay_medium)
        except Exception:
            pass

        return True
    except NoSuchElementException:
        print("  ✗ Could not find Save button!")
        return False


# ── Pay All Saved Installments + Reference Number capture ──

def click_pay_and_get_reference(driver, wait):
    """
    On the 'Selected Recurring Deposit Account List' page:
    1. Click 'Pay All Saved Installments'
    2. Parse the reference number from the success message
    Returns (success, reference_id).
    """
    # Wait for the saved list page
    try:
        wait.until(EC.presence_of_element_located(
            (By.XPATH, "//*[contains(text(), 'Selected Recurring Deposit Account List')]")
        ))
        print("  ✓ On 'Selected Recurring Deposit Account List' page")
    except TimeoutException:
        print("  ⚠ Could not confirm saved list page, trying anyway...")

    time.sleep(dashboard_state.delay_short)

    # Click "Pay All Saved Installments"
    try:
        pay_btn = driver.find_element(By.XPATH,
            "//input[@value='Pay All Saved Installments' or contains(@value,'Pay All Saved')]")
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", pay_btn)
        time.sleep(dashboard_state.delay_short)
        pay_btn.click()
        print("  ✓ Clicked 'Pay All Saved Installments'")
        time.sleep(dashboard_state.delay_long)

        # Handle any confirmation alert
        try:
            alert = driver.switch_to.alert
            print(f"    Alert: {alert.text}")
            alert.accept()
            time.sleep(dashboard_state.delay_medium)
        except Exception:
            pass

    except NoSuchElementException:
        print("  ✗ Could not find 'Pay All Saved Installments' button!")
        return False, ""

    # Wait for success message and extract reference number
    # Message: "Payment successful. Your payment reference number is C320461082."
    time.sleep(dashboard_state.delay_medium)
    reference_id = ""
    try:
        success_el = driver.find_element(By.XPATH,
            "//*[contains(text(), 'Payment successful') or contains(text(), 'payment reference')]")
        msg_text = success_el.text.strip()
        print(f"  ✓ {msg_text}")

        # Extract reference number
        m = re.search(r'reference\s+number\s+is\s+([A-Za-z0-9]+)', msg_text)
        if m:
            reference_id = m.group(1)
            print(f"  ✓ Reference ID captured: {reference_id}")
        else:
            # Try broader pattern (e.g. "C320461082" standalone)
            m = re.search(r'([A-Z]\d{6,})', msg_text)
            if m:
                reference_id = m.group(1)
                print(f"  ✓ Reference ID captured: {reference_id}")
            else:
                print(f"  ⚠ Could not parse reference ID from: '{msg_text}'")
                # Auto-pause so the user can read the page and enter via dashboard log
                with dashboard_state.lock:
                    dashboard_state.is_paused = True
                    dashboard_state.current_step = "PAUSED: Could not parse Reference ID — check portal, then Resume"
                control_flags.pause_event.clear()
                control_flags.pause_event.wait()
                # After resume, try one more scrape of the page
                try:
                    el2 = driver.find_element(By.XPATH,
                        "//*[contains(text(), 'reference number') or contains(text(), 'Reference')]")
                    m2 = re.search(r'([A-Z]\d{6,})', el2.text)
                    reference_id = m2.group(1) if m2 else ""
                    if reference_id:
                        print(f"  ✓ Reference ID captured after resume: {reference_id}")
                    else:
                        print(f"  ⚠ Still could not parse reference ID — Pay_Status will be FAIL")
                        return False, ""
                except Exception:
                    print(f"  ⚠ Reference ID not found — Pay_Status will be FAIL")
                    return False, ""
    except NoSuchElementException:
        print("  ⚠ Could not find success message on page")
        # Targeted fallback: search for any element mentioning 'reference'
        try:
            ref_el = driver.find_element(By.XPATH,
                "//*[contains(text(), 'reference number') or contains(text(), 'Reference')]")
            ref_text = ref_el.text.strip()
            m = re.search(r'reference\s+number\s+is\s+([A-Za-z0-9]+)', ref_text)
            if m:
                reference_id = m.group(1)
                print(f"  ✓ Reference ID from page: {reference_id}")
            else:
                m = re.search(r'([A-Z]\d{6,})', ref_text)
                if m:
                    reference_id = m.group(1)
                    print(f"  ✓ Reference ID from page: {reference_id}")
                else:
                    print(f"  ⚠ Reference ID not found in fallback element — Pay_Status will be FAIL")
                    return False, ""
        except Exception:
            print(f"  ⚠ No success or reference element found on page — Pay_Status will be FAIL")
            return False, ""

    return True, reference_id


# ── Phase 2: Reports → Download PDFs ──

def navigate_to_reports(driver, wait):
    """Click 'Reports' in the left sidebar to go to Recurring Deposit Installment Report."""
    try:
        reports_link = driver.find_element(By.XPATH, "//a[contains(text(), 'Reports')]")
        reports_link.click()
        time.sleep(dashboard_state.delay_medium)
        # Verify we're on the reports page
        wait.until(EC.presence_of_element_located(
            (By.XPATH, "//*[contains(text(), 'RECURRING DEPOSIT INSTALLMENT REPORT')]")
        ))
        print("✓ On Recurring Deposit Installment Report page")
        return True
    except (NoSuchElementException, TimeoutException):
        print("✗ Could not navigate to Reports page")
        input("  Please navigate to Reports manually, then press ENTER...")
        return True


def find_reference_input(driver):
    """Locate the 'List Reference No' input field on the Reports page."""
    # Try by name attribute
    try:
        return driver.find_element(By.XPATH,
            "//input[contains(@name, 'referenceNo') or contains(@name, 'Reference') or contains(@name, 'listRef')]")
    except NoSuchElementException:
        pass
    # Try by label proximity
    try:
        label = driver.find_element(By.XPATH, "//*[contains(text(), 'List Reference No')]")
        parent = label.find_element(By.XPATH, "./..")
        return parent.find_element(By.XPATH, ".//input[@type='text']")
    except NoSuchElementException:
        pass
    # Last resort: find all text inputs, skip date and cheque fields
    inputs = driver.find_elements(By.XPATH, "//input[@type='text']")
    for inp in inputs:
        val = inp.get_attribute("value") or ""
        name = (inp.get_attribute("name") or "").lower()
        if "date" in name or re.match(r'\d{2}-\w{3}-\d{4}', val):
            continue
        if "cheque" in name:
            continue
        return inp
    return None


def search_by_reference(driver, wait, reference_id):
    """Clear old reference, type new one, and click Search."""
    ref_input = find_reference_input(driver)
    if not ref_input:
        print(f"  ✗ Could not find List Reference No input!")
        return False

    # Clear old value manually (select all + delete) — avoids clicking Clear button
    ref_input.click()
    time.sleep(dashboard_state.delay_checkbox)
    ref_input.send_keys(SELECT_ALL_KEY, "a")
    time.sleep(dashboard_state.delay_checkbox)
    ref_input.send_keys(Keys.DELETE)
    time.sleep(dashboard_state.delay_checkbox)

    # Type new reference ID
    ref_input.send_keys(reference_id)
    time.sleep(dashboard_state.delay_short)
    print(f"  ✓ Entered reference: {reference_id}")

    # Click Search
    try:
        search_btn = driver.find_element(By.XPATH, "//input[@value='Search']")
        search_btn.click()
        print(f"  ✓ Clicked Search")
        time.sleep(dashboard_state.delay_medium)
        # Wait for results
        try:
            wait.until(EC.presence_of_element_located(
                (By.XPATH, "//*[contains(text(), 'Displaying')]")
            ))
        except TimeoutException:
            time.sleep(dashboard_state.delay_medium)
        return True
    except NoSuchElementException:
        print(f"  ✗ Could not find Search button!")
        return False


def wait_for_download(download_dir, timeout=30):
    """
    Wait for a new PDF to finish downloading in download_dir.
    Returns the path to the downloaded file, or None on timeout.
    Waits for: (1) .crdownload to disappear, (2) file size to stabilise.
    """
    end_time = time.time() + timeout
    while time.time() < end_time:
        crdownloads = glob_mod.glob(os.path.join(download_dir, "*.crdownload"))
        files = glob_mod.glob(os.path.join(download_dir, "*.pdf"))

        if files and not crdownloads:
            latest = max(files, key=os.path.getmtime)
            # Confirm size is stable (not still being flushed to disk)
            try:
                size1 = os.path.getsize(latest)
                time.sleep(0.5)
                size2 = os.path.getsize(latest)
                if size1 == size2 and size1 > 0:
                    return latest
            except OSError:
                pass  # File disappeared — keep waiting

        time.sleep(1)

    return None


def download_pdf_for_lot(driver, wait, lot_data, download_dir):
    """
    Search for a LOT's reference ID, verify count, and download PDF.
    Returns True on success.
    """
    lot = lot_data["LOT"]
    ref_id = lot_data.get("Reference_ID", "")
    expected_count = lot_data["Count"]

    if not ref_id:
        print(f"  ⚠ LOT {lot}: No Reference ID, skipping download")
        return False

    print(f"\n  {'─' * 50}")
    print(f"  Downloading LOT {lot}  |  Ref: {ref_id}  |  Count: {expected_count}")
    print(f"  {'─' * 50}")

    # Get existing PDF files before this download (to identify the new one)
    existing_pdfs = set(glob_mod.glob(os.path.join(download_dir, "*.pdf")))

    # Search by reference
    if not search_by_reference(driver, wait, ref_id):
        return False

    time.sleep(dashboard_state.delay_short)

    # Verify count
    display_text = get_display_text(driver)
    parsed = parse_display_count(display_text)
    if parsed:
        _, _, total = parsed
        print(f"  Site: '{display_text}'")
        if total == expected_count:
            print(f"  ✓ Count matches: {total}/{expected_count}")
        else:
            print(f"  ⚠ Count mismatch: site={total}, csv={expected_count}")
    else:
        print(f"  ⚠ Could not read display count")

    time.sleep(dashboard_state.delay_short)

    # Ensure "PDF file" is selected in the dropdown
    try:
        from selenium.webdriver.support.ui import Select
        dropdown = driver.find_element(By.XPATH, "//select[contains(@name,'download') or contains(@name,'format')]")
        select = Select(dropdown)
        select.select_by_visible_text("PDF file")
        time.sleep(0.5)
    except (NoSuchElementException, Exception):
        # May already be selected as PDF
        pass

    # Click OK to download
    try:
        ok_btn = driver.find_element(By.XPATH, "//input[@value='OK']")
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", ok_btn)
        time.sleep(dashboard_state.delay_short)
        ok_btn.click()
        print(f"  ✓ Clicked OK to download PDF")
        time.sleep(dashboard_state.delay_medium)

        # Handle any alert
        try:
            alert = driver.switch_to.alert
            print(f"    Alert: {alert.text}")
            alert.accept()
            time.sleep(dashboard_state.delay_short)
        except Exception:
            pass

    except NoSuchElementException:
        print(f"  ✗ Could not find OK button!")
        return False

    # Wait for the download to complete
    print(f"  Waiting for download...")
    time.sleep(dashboard_state.delay_medium)

    # Find the newly downloaded file
    new_pdfs = set(glob_mod.glob(os.path.join(download_dir, "*.pdf")))
    new_files = new_pdfs - existing_pdfs

    if not new_files:
        # Wait a bit more for the download
        downloaded = wait_for_download(download_dir, timeout=20)
        if downloaded and downloaded not in existing_pdfs:
            new_files = {downloaded}

    if new_files:
        downloaded_file = max(new_files, key=os.path.getmtime)
        # Rename to LOT#_RefID.pdf
        target_name = f"{lot}_{ref_id}.pdf"
        target_path = os.path.join(download_dir, target_name)
        try:
            os.rename(downloaded_file, target_path)
            print(f"  ✓ Saved as: {target_name}")
        except OSError as e:
            print(f"  ⚠ Could not rename file: {e}")
            print(f"    Downloaded to: {downloaded_file}")
    else:
        print(f"  ⚠ Could not detect downloaded file")
        print(f"    Check {download_dir} manually")

    time.sleep(dashboard_state.delay_short)

    # Do NOT click Clear button (it resets all fields including dates).
    # The reference field will be cleared manually in search_by_reference() next time.

    return True


def run_phase2(driver, wait, lots, download_dir):
    """Phase 2: Navigate to Reports and download PDFs for all completed LOTs."""
    print(f"\n{'=' * 60}")
    print(f"  PHASE 2: DOWNLOADING PDFs")
    print(f"  Download folder: {download_dir}")
    print(f"{'=' * 60}")

    os.makedirs(download_dir, exist_ok=True)

    # Navigate to Reports
    navigate_to_reports(driver, wait)
    time.sleep(dashboard_state.delay_short)

    download_success = 0
    download_fail = 0

    for lot_data in lots:
        ref_id = lot_data.get("Reference_ID", "")
        if not ref_id:
            continue  # Skip LOTs without reference IDs

        # Check if already downloaded
        target_name = f"{lot_data['LOT']}_{ref_id}.pdf"
        target_path = os.path.join(download_dir, target_name)
        if os.path.exists(target_path):
            print(f"\n  LOT {lot_data['LOT']}: {target_name} already exists, skipping.")
            download_success += 1
            continue

        try:
            result = download_pdf_for_lot(driver, wait, lot_data, download_dir)
            if result:
                download_success += 1
            else:
                download_fail += 1
        except Exception as e:
            print(f"\n  ✗ Error downloading LOT {lot_data['LOT']}: {e}")
            download_fail += 1

        time.sleep(dashboard_state.delay_long)

    print(f"\n  Phase 2 Summary: {download_success} downloaded, {download_fail} failed")
    print(f"  Files in: {download_dir}")


# ── Phase 3: Merge single-page PDFs ──

def _format_lot_range(lot_numbers):
    """Format LOT numbers as compact range string. e.g. [1,2,3,5,7,8,9] → '1-3,5,7-9'."""
    if not lot_numbers:
        return ""
    nums = sorted(lot_numbers)
    ranges = []
    start = nums[0]
    prev = nums[0]
    for n in nums[1:]:
        if n == prev + 1:
            prev = n
        else:
            ranges.append(f"{start}-{prev}" if start != prev else str(start))
            start = n
            prev = n
    ranges.append(f"{start}-{prev}" if start != prev else str(start))
    return ",".join(ranges)


def merge_single_page_pdfs(download_dir, lots):
    """Merge only single-page PDFs into one file. Multi-page PDFs are skipped."""
    try:
        from pypdf import PdfReader, PdfWriter
    except ImportError:
        print("  ⚠ pypdf not installed, skipping PDF merge (pip install pypdf)")
        return

    print(f"\n{'=' * 60}")
    print(f"  PHASE 3: MERGING SINGLE-PAGE PDFs")
    print(f"{'=' * 60}")

    # Collect LOT PDFs that exist on disk
    merged_lots = []
    skipped_lots = []
    writer = PdfWriter()

    for lot_data in lots:
        ref_id = lot_data.get("Reference_ID", "")
        if not ref_id:
            continue
        lot_num = lot_data["LOT"]
        target_name = f"{lot_num}_{ref_id}.pdf"
        target_path = os.path.join(download_dir, target_name)

        if not os.path.exists(target_path):
            continue

        try:
            reader = PdfReader(target_path)
            page_count = len(reader.pages)
            if page_count == 1:
                writer.add_page(reader.pages[0])
                merged_lots.append(int(lot_num) if lot_num.isdigit() else lot_num)
                print(f"  ✓ LOT {lot_num}: 1 page → included")
            else:
                skipped_lots.append((lot_num, page_count))
                print(f"  ✗ LOT {lot_num}: {page_count} pages → skipped")
        except Exception as e:
            skipped_lots.append((lot_num, f"error: {e}"))
            print(f"  ✗ LOT {lot_num}: could not read → skipped ({e})")

    if len(merged_lots) < 2:
        print(f"\n  Only {len(merged_lots)} single-page PDF(s) found, nothing to merge.")
        return

    # Build filename from LOT numbers
    if all(isinstance(n, int) for n in merged_lots):
        range_str = _format_lot_range(merged_lots)
    else:
        range_str = ",".join(str(n) for n in merged_lots)
    merged_filename = f"Merged_{range_str}.pdf"
    merged_path = os.path.join(download_dir, merged_filename)

    if os.path.exists(merged_path):
        print(f"\n  {merged_filename} already exists, skipping merge.")
        return

    try:
        writer.write(merged_path)
        print(f"\n  ✓ Merged PDF saved: {merged_filename}")
    except Exception as e:
        print(f"\n  ✗ Could not write merged PDF: {e}")
        return

    # Summary
    print(f"\n  {'─' * 50}")
    print(f"  MERGE SUMMARY")
    print(f"  {'─' * 50}")
    if all(isinstance(n, int) for n in merged_lots):
        print(f"  Merged LOTs : {_format_lot_range(merged_lots)}  ({len(merged_lots)} files)")
    else:
        print(f"  Merged LOTs : {', '.join(str(n) for n in merged_lots)}  ({len(merged_lots)} files)")
    if skipped_lots:
        for lot_num, reason in skipped_lots:
            print(f"  Skipped LOT {lot_num}: {reason} page(s)")
    print(f"  Output      : {merged_path}")


# ── Main LOT processing ──

def process_lot(driver, wait, lot_data, is_first_lot):
    """
    Process a single LOT. Updates lot_data dict in-place with status columns.
    Returns True on success, False on skip/failure.
    Raises SkipLotException if user skips via dashboard.
    """
    lot = lot_data["LOT"]
    rd_numbers = lot_data["RD Numbers"]
    expected_count = lot_data["Count"]
    remarks = []

    print(f"\n{'─' * 60}")
    print(f"  LOT {lot}  |  Expected: {expected_count} accounts")
    print(f"  RD Numbers: {rd_numbers[:80]}{'...' if len(rd_numbers) > 80 else ''}")
    print(f"{'─' * 60}")

    # ── Step 1: Ensure Cash mode ──
    checkpoint(dashboard_state, control_flags, "Step 1: Ensuring Cash mode")
    ensure_cash_mode(driver)
    time.sleep(dashboard_state.delay_short)

    # ── Step 2: Clear textarea and enter account numbers ──
    checkpoint(dashboard_state, control_flags, "Step 2: Entering RD numbers")
    if is_first_lot:
        try:
            textarea = driver.find_element(By.TAG_NAME, "textarea")
            current_value = textarea.get_attribute("value") or ""
            if current_value.strip():
                clear_btn = driver.find_element(By.XPATH,
                    "//input[@value='Clear Account' or contains(@value,'Clear')]")
                clear_btn.click()
                time.sleep(dashboard_state.delay_medium)
                ensure_cash_mode(driver)
                time.sleep(dashboard_state.delay_short)
        except (NoSuchElementException, Exception):
            pass

    if not clear_textarea_and_enter(driver, rd_numbers):
        lot_data["Fetch_Status"] = "FAIL"
        lot_data["Remarks"] = "Could not enter account IDs"
        return False
    time.sleep(dashboard_state.delay_short)

    # ── Step 3: Click Fetch ──
    checkpoint(dashboard_state, control_flags, "Step 3: Clicking Fetch")
    if not click_fetch(driver, wait):
        lot_data["Fetch_Status"] = "FAIL"
        lot_data["Remarks"] = "Fetch button not found"
        return False

    lot_data["Fetch_Status"] = "OK"
    time.sleep(dashboard_state.delay_short)

    # ── Step 4: Verify count ──
    checkpoint(dashboard_state, control_flags, "Step 4: Verifying count")
    display_text = get_display_text(driver)
    parsed = parse_display_count(display_text)

    if parsed:
        start, end, total = parsed
        print(f"  Site: '{display_text}'")
        print(f"  Fetched total: {total}  |  Expected: {expected_count}")

        if total == expected_count:
            lot_data["Count_Match"] = f"OK ({total}/{expected_count})"
            print(f"  ✓ Count MATCHES!")
        else:
            lot_data["Count_Match"] = f"MISMATCH ({total}/{expected_count})"
            print(f"  ⚠ Count MISMATCH! Auto-pausing for review...")
            remarks.append(f"Count mismatch: site={total} csv={expected_count}")
            # Auto-pause so user can decide via dashboard
            control_flags.pause_event.clear()
            with dashboard_state.lock:
                dashboard_state.is_paused = True
                dashboard_state.current_step = "PAUSED: Count mismatch - Resume to continue, Skip to skip LOT"
            control_flags.pause_event.wait()  # Blocks until user resumes or skips
            if control_flags.skip_lot.is_set():
                control_flags.skip_lot.clear()
                raise SkipLotException()
    else:
        lot_data["Count_Match"] = "UNREADABLE"
        print(f"  ⚠ Could not read display count, auto-pausing...")
        control_flags.pause_event.clear()
        with dashboard_state.lock:
            dashboard_state.is_paused = True
            dashboard_state.current_step = "PAUSED: Unreadable count - verify manually, then Resume"
        control_flags.pause_event.wait()
        if control_flags.skip_lot.is_set():
            control_flags.skip_lot.clear()
            raise SkipLotException()

    time.sleep(dashboard_state.delay_short)

    # ── Step 5: Validate due dates ──
    checkpoint(dashboard_state, control_flags, "Step 5: Validating due dates")
    print(f"  Checking due dates (expecting: {CURRENT_MONTH_ABBR})...")
    bad_rows = validate_due_dates_all_pages(driver, expected_count)

    if bad_rows:
        lot_data["Due_Date_Check"] = f"FAIL ({len(bad_rows)} bad)"
        print(f"\n  ✗ Due date mismatch in LOT {lot}!")
        for acct, due in bad_rows:
            print(f"    {acct}  →  {due}")
        remarks.append(f"Due date mismatch: {', '.join(a for a,d in bad_rows)}")
        lot_data["Remarks"] = "; ".join(remarks)
        print(f"  ⚠ Skipping LOT {lot}")
        return False
    else:
        lot_data["Due_Date_Check"] = "OK"
        print(f"  ✓ All due dates in {CURRENT_MONTH_ABBR}")

    time.sleep(dashboard_state.delay_short)

    # ── Step 6: Select all checkboxes ──
    checkpoint(dashboard_state, control_flags, "Step 6: Selecting checkboxes")
    print(f"  Selecting all checkboxes...")
    total_selected = select_all_checkboxes_all_pages(driver, expected_count)
    lot_data["Selected"] = str(total_selected)
    print(f"  Total selected: {total_selected}")

    time.sleep(dashboard_state.delay_short)

    # ── Step 7: Verify selection count ──
    checkpoint(dashboard_state, control_flags, "Step 7: Verifying selection")
    display_text_after = get_display_text(driver)
    parsed_after = parse_display_count(display_text_after)

    if parsed_after:
        s2, e2, t2 = parsed_after
        print(f"  After selection: '{display_text_after}'")
        if total_selected == t2:
            lot_data["Selection_Verified"] = f"OK ({total_selected}/{t2})"
            print(f"  ✓ Selection verified: {total_selected} of {t2}")
        else:
            lot_data["Selection_Verified"] = f"MISMATCH ({total_selected}/{t2})"
            print(f"  ⚠ Selection mismatch: selected={total_selected}, total={t2}")
            remarks.append(f"Selection mismatch: {total_selected}/{t2}")
            # Auto-pause for review
            control_flags.pause_event.clear()
            with dashboard_state.lock:
                dashboard_state.is_paused = True
                dashboard_state.current_step = "PAUSED: Selection mismatch - Resume to save, Skip to skip LOT"
            control_flags.pause_event.wait()
            if control_flags.skip_lot.is_set():
                control_flags.skip_lot.clear()
                raise SkipLotException()
    else:
        if total_selected == expected_count:
            lot_data["Selection_Verified"] = f"OK ({total_selected}/{expected_count})"
            print(f"  ✓ Selected {total_selected} = expected {expected_count}")
        else:
            lot_data["Selection_Verified"] = f"CHECK ({total_selected}/{expected_count})"
            print(f"  ⚠ Selected {total_selected}, expected {expected_count}")
            remarks.append(f"Selection check: {total_selected}/{expected_count}")

    time.sleep(dashboard_state.delay_short)

    # ── Step 8: Click Save ──
    checkpoint(dashboard_state, control_flags, "Step 8: Saving")
    if not click_save(driver, wait):
        lot_data["Save_Status"] = "FAIL"
        remarks.append("Save button not found")
        lot_data["Remarks"] = "; ".join(remarks)
        return False

    lot_data["Save_Status"] = "OK"
    time.sleep(dashboard_state.delay_short)

    # ── Step 9: Pay All Saved Installments ──
    checkpoint(dashboard_state, control_flags, "Step 9: Paying installments")
    print(f"  Paying All Saved Installments...")
    pay_ok, ref_id = click_pay_and_get_reference(driver, wait)

    if not pay_ok:
        lot_data["Pay_Status"] = "FAIL"
        remarks.append("Pay All Saved Installments failed")
        lot_data["Remarks"] = "; ".join(remarks)
        return False

    lot_data["Pay_Status"] = "OK"
    lot_data["Reference_ID"] = ref_id
    lot_data["Remarks"] = "; ".join(remarks) if remarks else "Success"

    print(f"  ✓ LOT {lot} fully completed! Ref: {ref_id}")

    # After Pay, we're back on the Deposit Accounts page automatically
    time.sleep(dashboard_state.delay_short)

    return True


# ── Entry point ──

def main():
    global XLSX_FILE

    # ── File path prompt ──
    print("\n" + "=" * 60)
    print("  DOP AUTOMATION — SELECT INPUT FILE")
    print("=" * 60)
    print("\n  Tip (macOS): drag-and-drop your .xlsx file into this")
    print("  terminal window to auto-fill its path, then press Enter.\n")

    while True:
        raw = input("  Excel file path (.xlsx): ").strip().strip('"').strip("'")
        if not raw:
            print("  ⚠ No path entered, try again.")
            continue
        if not os.path.isfile(raw):
            print(f"  ✗ File not found: {raw}")
            continue
        if not raw.lower().endswith(".xlsx"):
            print(f"  ✗ Must be an .xlsx file (got: {os.path.basename(raw)})")
            continue
        XLSX_FILE = raw
        break

    print(f"\n  File: {XLSX_FILE}")

    # Read Excel
    lots = read_xlsx(XLSX_FILE)
    start_global_timeout(lots)
    print(f"\nLoaded {len(lots)} LOTs from Excel")
    print(f"Current month for validation: {CURRENT_MONTH_ABBR}")
    print()
    for lot in lots:
        pay = lot.get("Pay_Status", "")
        save = lot.get("Save_Status", "")
        ref = lot.get("Reference_ID", "")
        if pay == "OK":
            marker = f" (done - Ref: {ref})" if ref else " (done)"
        elif save == "OK":
            marker = " (saved, NOT yet paid)"
        else:
            marker = ""
        print(f"  LOT {lot['LOT']}: {lot['Count']} accounts{marker}")

    # Ask which LOTs to process (terminal input, before automation starts)
    print(f"\nProcess all LOTs (1-{len(lots)})? Or specify range.")
    range_input = input("Enter 'all' or range like '1-5' or specific like '1,3,5': ").strip()

    try:
        if range_input.lower() == 'all' or range_input == '':
            lots_to_process = list(range(len(lots)))
        elif '-' in range_input and ',' not in range_input:
            parts = range_input.split('-', 1)
            start_idx = max(0, int(parts[0].strip()) - 1)
            end_idx   = min(len(lots), int(parts[1].strip()))
            if start_idx >= end_idx:
                print(f"  ⚠ Invalid range '{range_input}', processing all LOTs.")
                lots_to_process = list(range(len(lots)))
            else:
                lots_to_process = list(range(start_idx, end_idx))
        else:
            lots_to_process = []
            for x in range_input.split(','):
                idx = int(x.strip()) - 1
                if 0 <= idx < len(lots):
                    lots_to_process.append(idx)
                else:
                    print(f"  ⚠ LOT {x.strip()} out of range (1–{len(lots)}), skipping.")
        if not lots_to_process:
            print("  No valid LOTs selected, defaulting to all.")
            lots_to_process = list(range(len(lots)))
    except ValueError:
        print("  Invalid input, processing all LOTs.")
        lots_to_process = list(range(len(lots)))

    # Initialize dashboard state
    dashboard_state.start_time = time.time()
    dashboard_state.lots_total = len(lots_to_process)
    dashboard_state.current_phase = "Phase 1"
    dashboard_state.delay_short = DELAY_SHORT
    dashboard_state.delay_medium = DELAY_MEDIUM
    dashboard_state.delay_long = DELAY_LONG
    dashboard_state.delay_checkbox = DELAY_CHECKBOX
    dashboard_state.lot_statuses = [
        {
            "lot": lots[i]["LOT"],
            "count": lots[i]["Count"],
            "status": "done" if lots[i].get("Pay_Status") == "OK" else "pending",
            "ref_id": lots[i].get("Reference_ID", ""),
            "step": ""
        }
        for i in range(len(lots))
    ]

    # Start web dashboard
    start_dashboard(dashboard_state, control_flags)
    print(f"\n  Open the dashboard URL above in your browser for live control.\n")

    # Setup browser (with download dir set for Phase 2)
    os.makedirs(DOWNLOAD_DIR, exist_ok=True)
    driver = setup_driver(download_dir=DOWNLOAD_DIR)
    wait = WebDriverWait(driver, WAIT_TIMEOUT)

    # Wait for manual login
    with dashboard_state.lock:
        dashboard_state.current_step = "Waiting for manual login..."
    wait_for_login(driver)
    navigate_to_deposit_accounts(driver, wait)

    # Process each LOT
    success_count = 0
    fail_count = 0
    skip_count = 0
    is_first_lot = True

    for idx in lots_to_process:
        lot_data = lots[idx]

        # Skip if already fully done
        if lot_data.get("Pay_Status") == "OK":
            ref = lot_data.get("Reference_ID", "")
            print(f"\n  LOT {lot_data['LOT']} already done (Ref: {ref}), skipping.")
            continue

        # Check if user marked this LOT to skip via dashboard
        with control_flags.lock:
            if lot_data["LOT"] in control_flags.skip_lots_set:
                print(f"\n  LOT {lot_data['LOT']} skipped by user (dashboard).")
                with dashboard_state.lock:
                    dashboard_state.lot_statuses[idx]["status"] = "skipped"
                    dashboard_state.lots_skipped += 1
                skip_count += 1
                continue

        # Update dashboard state
        with dashboard_state.lock:
            dashboard_state.current_lot = lot_data["LOT"]
            dashboard_state.lot_statuses[idx]["status"] = "running"
            dashboard_state.current_step = "Starting LOT"

        # Memory watchdog
        mem_ok = check_memory_usage(driver)
        with dashboard_state.lock:
            dashboard_state.memory_mb = _get_memory_mb()
        if not mem_ok:
            write_xlsx(XLSX_FILE, lots)
            print("  Progress saved. Restart the script to continue (it will resume).")
            sys.exit(1)

        try:
            result = process_lot(driver, wait, lot_data, is_first_lot)
            is_first_lot = False
            if result:
                success_count += 1
                with dashboard_state.lock:
                    dashboard_state.lot_statuses[idx]["status"] = "done"
                    dashboard_state.lot_statuses[idx]["ref_id"] = lot_data.get("Reference_ID", "")
                    dashboard_state.lots_done += 1
            else:
                fail_count += 1
                with dashboard_state.lock:
                    dashboard_state.lot_statuses[idx]["status"] = "failed"
                    dashboard_state.lots_failed += 1

        except SkipLotException:
            print(f"\n  LOT {lot_data['LOT']} skipped by user.")
            lot_data["Remarks"] = "Skipped by user"
            lot_data["Pay_Status"] = "SKIP"
            skip_count += 1
            is_first_lot = False
            with dashboard_state.lock:
                dashboard_state.lot_statuses[idx]["status"] = "skipped"
                dashboard_state.lots_skipped += 1

        except StopAfterCurrentException:
            print(f"\n  Stop requested by user. Saving and exiting.")
            write_xlsx(XLSX_FILE, lots)
            print(f"  Progress saved.")
            break

        except Exception as e:
            print(f"\n✗ Error processing LOT {lot_data['LOT']}: {e}")
            lot_data["Remarks"] = f"Error: {e}"
            fail_count += 1
            is_first_lot = False
            with dashboard_state.lock:
                dashboard_state.lot_statuses[idx]["status"] = "failed"
                dashboard_state.lots_failed += 1
            # Auto-continue instead of blocking on input()
            print("  Continuing to next LOT automatically...")

        # Save progress after every LOT
        write_xlsx(XLSX_FILE, lots)
        print(f"  Progress saved to XLSX")

        # Breathing room between LOTs
        time.sleep(dashboard_state.delay_long)

    # Final save
    write_xlsx(XLSX_FILE, lots)

    # Phase 1 Summary
    with dashboard_state.lock:
        dashboard_state.current_phase = "Phase 1 Complete"
        dashboard_state.current_step = "Waiting for Phase 2 decision..."

    print(f"\n{'=' * 60}")
    print(f"  PHASE 1 SUMMARY")
    print(f"  Successful : {success_count}")
    print(f"  Failed     : {fail_count}")
    print(f"  Skipped    : {skip_count}")
    print(f"  Total      : {success_count + fail_count + skip_count}")
    print(f"{'=' * 60}")
    print(f"  XLSX : {XLSX_FILE} (Reference_ID column in GREEN)")

    # Check if any LOTs have reference IDs for Phase 2
    lots_with_refs = [l for l in lots if l.get("Reference_ID", "")]
    if lots_with_refs:
        print(f"\n  {len(lots_with_refs)} LOTs have Reference IDs ready for PDF download.")
        phase2_input = input("\n  Start Phase 2 (PDF Downloads)? (y/n): ").strip().lower()
        if phase2_input == 'y':
            with dashboard_state.lock:
                dashboard_state.current_phase = "Phase 2"
                dashboard_state.current_step = "Downloading PDFs..."
            run_phase2(driver, wait, lots, DOWNLOAD_DIR)

            with dashboard_state.lock:
                dashboard_state.current_phase = "Phase 3"
                dashboard_state.current_step = "Merging PDFs..."
            merge_single_page_pdfs(DOWNLOAD_DIR, lots)
    else:
        print("\n  No LOTs with Reference IDs — skipping Phase 2.")

    with dashboard_state.lock:
        dashboard_state.is_finished = True
        dashboard_state.current_phase = "Complete"
        dashboard_state.current_step = "All done"

    print(f"\n  Browser will remain open. Close it manually when done.")


if __name__ == "__main__":
    main()
